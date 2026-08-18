import json
import openpyxl
from collections import Counter
from urllib.parse import urlparse
import glob
import re

# Load techbot.xlsx
wb = openpyxl.load_workbook('datasets/techbot.xlsx')
courses_sheet = wb['Courses']
readme_sheet = wb['README']

courses_rows = list(courses_sheet.iter_rows(values_only=True))
courses_header = courses_rows[0]
courses_data = [dict(zip(courses_header, r)) for r in courses_rows[1:]]

readme_rows = list(readme_sheet.iter_rows(values_only=True))
readme_data = [dict(zip(readme_rows[0], r)) for r in readme_rows[1:]]

# 1. Dataset Analysis
skills = sorted(list(set(r['skill_tag'] for r in courses_data)))
levels = sorted(list(set(r['level'] for r in courses_data)))
platforms = sorted(list(set(str(r['platform']) for r in courses_data)))
course_ids = [r['course_id'] for r in courses_data]
titles = [r['title'] for r in courses_data]
links = [str(r['link']) for r in courses_data]

# Duplicate analysis
dup_course_ids = [cid for cid, count in Counter(course_ids).items() if count > 1]
title_counts = Counter(titles)
dup_titles = {t: c for t, c in title_counts.items() if c > 1}
link_counts = Counter(links)
dup_links = {l: c for l, c in link_counts.items() if c > 1}

# Skill-Level combo uniqueness
skill_level_combos = [(r['skill_tag'], r['level']) for r in courses_data]
combo_counts = Counter(skill_level_combos)
dup_combos = {combo: c for combo, c in combo_counts.items() if c > 1}

# URL validation & YouTube check
youtube_links = []
malformed_urls = []
valid_urls = []

for r in courses_data:
    url = str(r['link']).strip()
    parsed = urlparse(url)
    if not (parsed.scheme in ['http', 'https'] and parsed.netloc):
        malformed_urls.append((r['course_id'], r['title'], url))
    else:
        valid_urls.append(url)
    if 'youtube.com' in url.lower() or 'youtu.be' in url.lower():
        youtube_links.append((r['course_id'], r['skill_tag'], r['level'], url))

# Check platforms
platform_counts = Counter(r['platform'] for r in courses_data)

# Print Summary
print("==================================================")
print("DATASET STATS")
print("==================================================")
print(f"Total Rows: {len(courses_data)}")
print(f"Unique Skills ({len(skills)}): {skills}")
print(f"Unique Levels ({len(levels)}): {levels}")
print(f"Duplicate Course IDs: {dup_course_ids}")
print(f"Duplicate Titles ({len(dup_titles)} distinct titles repeated across {sum(dup_titles.values())} rows):")
for t, c in list(dup_titles.items())[:10]:
    print(f"  '{t}': {c} occurrences")
print(f"Duplicate Links ({len(dup_links)} distinct links repeated across {sum(dup_links.values())} rows)")
print(f"Duplicate (Skill, Level) combos: {len(dup_combos)}")
print(f"Malformed URLs: {len(malformed_urls)}")
print(f"YouTube links: {len(youtube_links)}")
print(f"Top 10 Platforms: {platform_counts.most_common(10)}")

# Save dataset audit summary
audit_summary = {
    "total_courses": len(courses_data),
    "unique_skills_count": len(skills),
    "skills": skills,
    "unique_levels": levels,
    "level_counts": dict(Counter(r['level'] for r in courses_data)),
    "duration_hours_counts": dict(Counter(r['duration_hours'] for r in courses_data)),
    "unique_platforms_count": len(platforms),
    "top_platforms": platform_counts.most_common(15),
    "duplicate_titles_count": len(dup_titles),
    "duplicate_titles_sample": list(dup_titles.items())[:15],
    "duplicate_links_count": len(dup_links),
    "malformed_urls": malformed_urls,
    "youtube_links_count": len(youtube_links),
    "youtube_links_sample": youtube_links[:10],
    "readme": readme_data
}

with open('scratch/dataset_summary.json', 'w', encoding='utf-8') as f:
    json.dump(audit_summary, f, indent=2)

print("\nWrote summary to scratch/dataset_summary.json")
