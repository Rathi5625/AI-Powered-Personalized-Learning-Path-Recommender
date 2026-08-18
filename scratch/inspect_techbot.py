import openpyxl
import json
from collections import Counter
import re

wb = openpyxl.load_workbook('datasets/techbot.xlsx')
print(f"Worksheet names: {wb.sheetnames}")

full_audit = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print(f"Sheet {sheet_name} is empty")
        continue
    
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data_rows = rows[1:]
    
    print(f"\n--- Sheet: {sheet_name} ---")
    print(f"Total rows (incl header): {len(rows)}, Data rows: {len(data_rows)}")
    print(f"Headers: {headers}")
    
    sheet_data = []
    for r_idx, r in enumerate(data_rows, start=2):
        row_dict = {}
        for h, val in zip(headers, r):
            row_dict[h] = val
        sheet_data.append(row_dict)
    
    full_audit[sheet_name] = {
        "total_rows": len(rows),
        "data_rows_count": len(data_rows),
        "headers": headers,
        "sample_rows": sheet_data[:5],
        "all_rows": sheet_data
    }

# Let's inspect data characteristics
print("\n" + "="*50)
print("DETAILED DATA ANALYSIS")
print("="*50)

for sheet_name, s_info in full_audit.items():
    print(f"\nAnalyzing sheet: {sheet_name}")
    rows = s_info["all_rows"]
    headers = s_info["headers"]
    
    # Check null / missing values per column
    missing_counts = {h: sum(1 for r in rows if r.get(h) is None or str(r.get(h)).strip() == '') for h in headers}
    print("Missing values per column:", missing_counts)
    
    # Check distinct values for potential categoricals
    for h in headers:
        vals = [r.get(h) for r in rows if r.get(h) is not None]
        unique_vals = set(vals)
        print(f"Column '{h}': {len(unique_vals)} unique values (out of {len(vals)} non-empty)")
        if len(unique_vals) <= 20:
            print(f"  Values: {Counter(vals)}")

# Write full audit details to a json file for deep analysis
with open('datasets/techbot_audit_raw.json', 'w', encoding='utf-8') as f:
    json.dump(full_audit, f, indent=2, default=str)
print("\nWrote raw audit dump to datasets/techbot_audit_raw.json")
